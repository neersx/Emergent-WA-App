"""Analytics + usage + billing endpoints.

All reads are tenant-scoped. PlatformSuperAdmin cross-tenant reads live in platform.py.
"""
from __future__ import annotations

import csv
import io
import uuid
from datetime import datetime, timezone, timedelta
from typing import Any

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from loguru import logger

from .db import get_db
from .tenancy import Principal, require_tenant

router = APIRouter(prefix="/api/analytics", tags=["analytics"])

# ─── Date range helpers ───────────────────────────────────────────────────────

def _parse_range(days: int, from_date: str | None, to_date: str | None) -> tuple[str, str, str, str]:
    """Return (since_iso, since_day, to_day, to_iso) from params."""
    if from_date:
        since_dt = datetime.fromisoformat(from_date)
    else:
        since_dt = datetime.now(timezone.utc) - timedelta(days=days)
    if to_date:
        to_dt = datetime.fromisoformat(to_date) + timedelta(days=1)  # inclusive
    else:
        to_dt = datetime.now(timezone.utc) + timedelta(days=1)
    return (
        since_dt.isoformat(),
        since_dt.date().isoformat(),
        to_dt.date().isoformat(),
        to_dt.isoformat(),
    )


# ─── Country code extraction ──────────────────────────────────────────────────

_CC_PREFIXES: dict[str, str] = {
    "1": "US", "44": "GB", "49": "DE", "33": "FR", "34": "ES",
    "91": "IN", "55": "BR", "52": "MX", "62": "ID", "966": "SA",
    "971": "AE", "86": "CN", "81": "JP", "82": "KR", "27": "ZA",
    "234": "NG", "92": "PK", "63": "PH", "66": "TH", "84": "VN",
    "7": "RU", "90": "TR", "39": "IT", "31": "NL", "61": "AU",
    "46": "SE", "47": "NO", "45": "DK", "358": "FI", "48": "PL",
}


def wa_id_to_country(wa_id: str) -> str | None:
    if not wa_id:
        return None
    for prefix in sorted(_CC_PREFIXES, key=len, reverse=True):
        if wa_id.startswith(prefix):
            return _CC_PREFIXES[prefix]
    return None


# ─── Existing endpoints ───────────────────────────────────────────────────────

@router.get("/overview")
async def overview(
    p: Principal = Depends(require_tenant),
    days: int = Query(default=14, ge=1, le=90),
    from_date: str | None = Query(default=None),
    to_date: str | None = Query(default=None),
):
    db = get_db()
    since_iso, since_day, to_day, to_iso = _parse_range(days, from_date, to_date)

    base = {"tenant_id": p.tenant_id, "created_at": {"$gte": since_iso, "$lt": to_iso}}
    total = await db.messages.count_documents(base)
    outbound = await db.messages.count_documents({**base, "direction": "outbound"})
    inbound = await db.messages.count_documents({**base, "direction": "inbound"})

    status_pipeline = [
        {"$match": base},
        {"$group": {"_id": "$status", "count": {"$sum": 1}}},
    ]
    status_breakdown = {
        row["_id"] or "unknown": row["count"]
        async for row in db.messages.aggregate(status_pipeline)
    }

    delivered = status_breakdown.get("delivered", 0) + status_breakdown.get("read", 0)
    failed = status_breakdown.get("failed", 0)
    delivery_rate = delivered / outbound if outbound else 0.0
    read = status_breakdown.get("read", 0)
    read_rate = read / outbound if outbound else 0.0

    conv_total = await db.conversations.count_documents({"tenant_id": p.tenant_id})
    open_window_cutoff = (datetime.now(timezone.utc) - timedelta(hours=24)).isoformat()
    conv_open_window = await db.conversations.count_documents(
        {"tenant_id": p.tenant_id, "last_inbound_at": {"$gte": open_window_cutoff}}
    )

    return {
        "window_days": days,
        "from_date": since_day,
        "to_date": to_day,
        "total_messages": total,
        "outbound": outbound,
        "inbound": inbound,
        "status_breakdown": status_breakdown,
        "delivered": delivered,
        "read": read,
        "failed": failed,
        "delivery_rate": round(delivery_rate, 4),
        "read_rate": round(read_rate, 4),
        "conversations": conv_total,
        "conversations_open_window": conv_open_window,
    }


@router.get("/timeseries")
async def timeseries(
    p: Principal = Depends(require_tenant),
    days: int = Query(default=14, ge=1, le=90),
    from_date: str | None = Query(default=None),
    to_date: str | None = Query(default=None),
):
    db = get_db()
    since_iso, since_day, to_day, to_iso = _parse_range(days, from_date, to_date)

    # Build continuous day buckets
    since_dt = datetime.fromisoformat(since_day)
    to_dt = datetime.fromisoformat(to_day)
    diff_days = (to_dt - since_dt).days
    buckets: dict[str, dict] = {}
    for i in range(diff_days):
        d = (since_dt + timedelta(days=i)).date().isoformat()
        buckets[d] = {"date": d, "outbound": 0, "inbound": 0, "delivered": 0, "read": 0, "failed": 0}

    pipeline = [
        {"$match": {"tenant_id": p.tenant_id, "created_at": {"$gte": since_iso, "$lt": to_iso}}},
        {"$addFields": {"day": {"$substr": ["$created_at", 0, 10]}}},
        {"$group": {
            "_id": {"day": "$day", "direction": "$direction", "status": "$status"},
            "count": {"$sum": 1},
        }},
    ]
    async for row in db.messages.aggregate(pipeline):
        day = row["_id"]["day"]
        if day not in buckets:
            buckets[day] = {"date": day, "outbound": 0, "inbound": 0, "delivered": 0, "read": 0, "failed": 0}
        direction = row["_id"].get("direction") or ""
        status = row["_id"].get("status") or ""
        if direction == "outbound":
            buckets[day]["outbound"] += row["count"]
        elif direction == "inbound":
            buckets[day]["inbound"] += row["count"]
        if status in ("delivered", "read", "failed"):
            buckets[day][status] += row["count"]
    return sorted(buckets.values(), key=lambda b: b["date"])


@router.get("/by-template")
async def by_template(
    p: Principal = Depends(require_tenant),
    days: int = Query(default=14, ge=1, le=90),
    from_date: str | None = Query(default=None),
    to_date: str | None = Query(default=None),
    limit: int = Query(default=10, ge=1, le=50),
):
    db = get_db()
    since_iso, _, _, to_iso = _parse_range(days, from_date, to_date)
    pipeline = [
        {"$match": {
            "tenant_id": p.tenant_id,
            "created_at": {"$gte": since_iso, "$lt": to_iso},
            "direction": "outbound",
            "template_name": {"$ne": None},
        }},
        {"$group": {
            "_id": "$template_name",
            "sent": {"$sum": 1},
            "delivered": {"$sum": {"$cond": [{"$in": ["$status", ["delivered", "read"]]}, 1, 0]}},
            "read": {"$sum": {"$cond": [{"$eq": ["$status", "read"]}, 1, 0]}},
            "failed": {"$sum": {"$cond": [{"$eq": ["$status", "failed"]}, 1, 0]}},
        }},
        {"$sort": {"sent": -1}},
        {"$limit": limit},
    ]
    out = []
    async for r in db.messages.aggregate(pipeline):
        out.append({
            "template_name": r["_id"],
            "sent": r["sent"],
            "delivered": r["delivered"],
            "read": r["read"],
            "failed": r["failed"],
            "delivery_rate": round(r["delivered"] / r["sent"], 4) if r["sent"] else 0.0,
            "read_rate": round(r["read"] / r["sent"], 4) if r["sent"] else 0.0,
        })
    return out


@router.get("/by-phone")
async def by_phone(
    p: Principal = Depends(require_tenant),
    days: int = Query(default=14, ge=1, le=90),
    from_date: str | None = Query(default=None),
    to_date: str | None = Query(default=None),
):
    db = get_db()
    since_iso, _, _, to_iso = _parse_range(days, from_date, to_date)
    phones = {p2["phone_number_id"]: p2 async for p2 in db.phone_numbers.find({"tenant_id": p.tenant_id})}
    pipeline = [
        {"$match": {"tenant_id": p.tenant_id, "created_at": {"$gte": since_iso, "$lt": to_iso}}},
        {"$group": {
            "_id": "$phone_number_id",
            "outbound": {"$sum": {"$cond": [{"$eq": ["$direction", "outbound"]}, 1, 0]}},
            "inbound": {"$sum": {"$cond": [{"$eq": ["$direction", "inbound"]}, 1, 0]}},
        }},
    ]
    out = []
    async for r in db.messages.aggregate(pipeline):
        pid = r["_id"]
        phone = phones.get(pid)
        out.append({
            "phone_number_id": pid,
            "display": phone["display_phone_number"] if phone else pid,
            "verified_name": phone["verified_name"] if phone else None,
            "outbound": r["outbound"],
            "inbound": r["inbound"],
        })
    out.sort(key=lambda x: x["outbound"] + x["inbound"], reverse=True)
    return out


# ─── Core dashboard endpoint (comprehensive) ─────────────────────────────────

@router.get("/dashboard")
async def dashboard(
    p: Principal = Depends(require_tenant),
    days: int = Query(default=14, ge=1, le=90),
    from_date: str | None = Query(default=None),
    to_date: str | None = Query(default=None),
    phone_number_id: str | None = Query(default=None),
    waba_id: str | None = Query(default=None),
):
    """Comprehensive core dashboard: delivery/read/fail rates, cost burn, failure reasons, throughput."""
    db = get_db()
    since_iso, since_day, to_day, to_iso = _parse_range(days, from_date, to_date)

    base: dict[str, Any] = {
        "tenant_id": p.tenant_id,
        "created_at": {"$gte": since_iso, "$lt": to_iso},
    }
    if phone_number_id:
        base["phone_number_id"] = phone_number_id
    if waba_id:
        # Filter by phones in this WABA
        phone_ids = [
            d["phone_number_id"] async for d in
            db.phone_numbers.find({"tenant_id": p.tenant_id, "waba_id": waba_id})
        ]
        base["phone_number_id"] = {"$in": phone_ids}

    # Status counts
    status_pipeline = [
        {"$match": {**base, "direction": "outbound"}},
        {"$group": {"_id": "$status", "count": {"$sum": 1}}},
    ]
    status_bd: dict[str, int] = {}
    async for r in db.messages.aggregate(status_pipeline):
        status_bd[r["_id"] or "unknown"] = r["count"]

    outbound = sum(status_bd.values())
    delivered = status_bd.get("delivered", 0) + status_bd.get("read", 0)
    read_ = status_bd.get("read", 0)
    failed = status_bd.get("failed", 0)

    # Top failure reasons
    fail_pipeline = [
        {"$match": {**base, "direction": "outbound", "status": "failed", "error": {"$ne": None}}},
        {"$group": {"_id": "$error", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 5},
    ]
    fail_reasons = [{"reason": r["_id"], "count": r["count"]} async for r in db.messages.aggregate(fail_pipeline)]

    # Cost burn rate (daily from rollup)
    rollup_docs = await db.usage_daily_rollup.find(
        {"tenant_id": p.tenant_id, "day": {"$gte": since_day, "$lt": to_day}}
    ).sort("day", 1).to_list(500)

    # Aggregate cost by day for burn chart
    cost_by_day: dict[str, float] = {}
    for r in rollup_docs:
        cost_by_day[r["day"]] = cost_by_day.get(r["day"], 0.0) + r.get("cost_amount", 0.0)

    cost_series = [{"date": d, "cost": round(v, 4)} for d, v in sorted(cost_by_day.items())]
    total_cost = sum(v for v in cost_by_day.values())

    # Throughput (messages per day)
    throughput_pipeline = [
        {"$match": base},
        {"$addFields": {"day": {"$substr": ["$created_at", 0, 10]}}},
        {"$group": {"_id": "$day", "count": {"$sum": 1}}},
        {"$sort": {"_id": 1}},
    ]
    throughput = [{"date": r["_id"], "count": r["count"]} async for r in db.messages.aggregate(throughput_pipeline)]

    # Inbound
    inbound_count = await db.messages.count_documents({**base, "direction": "inbound"})

    return {
        "from_date": since_day,
        "to_date": to_day,
        "outbound": outbound,
        "inbound": inbound_count,
        "delivered": delivered,
        "read": read_,
        "failed": failed,
        "delivery_rate": round(delivered / outbound, 4) if outbound else 0.0,
        "read_rate": round(read_ / outbound, 4) if outbound else 0.0,
        "failure_rate": round(failed / outbound, 4) if outbound else 0.0,
        "status_breakdown": status_bd,
        "top_failure_reasons": fail_reasons,
        "cost_burn_rate": cost_series,
        "total_cost_usd": round(total_cost, 4),
        "throughput": throughput,
    }


# ─── Usage / cost endpoints ───────────────────────────────────────────────────

@router.get("/usage/daily")
async def usage_daily(
    p: Principal = Depends(require_tenant),
    days: int = Query(default=30, ge=1, le=365),
    from_date: str | None = Query(default=None),
    to_date: str | None = Query(default=None),
):
    db = get_db()
    _, since_day, to_day, _ = _parse_range(days, from_date, to_date)
    docs = await db.usage_daily_rollup.find(
        {"tenant_id": p.tenant_id, "day": {"$gte": since_day, "$lt": to_day}}
    ).sort("day", -1).to_list(500)
    return [
        {
            "day": d["day"],
            "category": d.get("category", "service"),
            "country_code": d.get("country_code"),
            "delivered_count": d.get("delivered_count", 0),
            "billable_count": d.get("billable_count", 0),
            "free_count": d.get("free_count", 0),
            "cost_amount": round(d.get("cost_amount", 0.0), 4),
            "cost_currency": d.get("cost_currency", "USD"),
        }
        for d in docs
    ]


@router.get("/usage/cost")
async def usage_cost_summary(
    p: Principal = Depends(require_tenant),
    days: int = Query(default=30, ge=1, le=365),
    from_date: str | None = Query(default=None),
    to_date: str | None = Query(default=None),
):
    db = get_db()
    _, since_day, to_day, _ = _parse_range(days, from_date, to_date)
    pipeline = [
        {"$match": {"tenant_id": p.tenant_id, "day": {"$gte": since_day, "$lt": to_day}}},
        {"$group": {
            "_id": "$category",
            "delivered_count": {"$sum": "$delivered_count"},
            "billable_count": {"$sum": "$billable_count"},
            "free_count": {"$sum": "$free_count"},
            "cost_amount": {"$sum": "$cost_amount"},
        }},
        {"$sort": {"cost_amount": -1}},
    ]
    by_category = []
    total_cost = 0.0
    total_delivered = 0
    async for r in db.usage_daily_rollup.aggregate(pipeline):
        cost = round(r.get("cost_amount", 0.0), 4)
        total_cost += cost
        total_delivered += r.get("delivered_count", 0)
        by_category.append({
            "category": r["_id"] or "service",
            "delivered_count": r.get("delivered_count", 0),
            "billable_count": r.get("billable_count", 0),
            "free_count": r.get("free_count", 0),
            "cost_amount": cost,
            "cost_currency": "USD",
        })

    # Cost by country
    country_pipeline = [
        {"$match": {"tenant_id": p.tenant_id, "day": {"$gte": since_day, "$lt": to_day}, "country_code": {"$ne": None}}},
        {"$group": {
            "_id": "$country_code",
            "delivered_count": {"$sum": "$delivered_count"},
            "cost_amount": {"$sum": "$cost_amount"},
        }},
        {"$sort": {"cost_amount": -1}},
        {"$limit": 20},
    ]
    by_country = [
        {
            "country_code": r["_id"],
            "delivered_count": r.get("delivered_count", 0),
            "cost_amount": round(r.get("cost_amount", 0.0), 4),
        }
        async for r in db.usage_daily_rollup.aggregate(country_pipeline)
    ]

    # MTD calculation
    mtd_start = datetime.now(timezone.utc).replace(day=1).date().isoformat()
    mtd_pipeline = [
        {"$match": {"tenant_id": p.tenant_id, "day": {"$gte": mtd_start}}},
        {"$group": {"_id": None, "cost": {"$sum": "$cost_amount"}, "delivered": {"$sum": "$delivered_count"}}},
    ]
    mtd_data = [d async for d in db.usage_daily_rollup.aggregate(mtd_pipeline)]
    mtd = mtd_data[0] if mtd_data else {"cost": 0, "delivered": 0}

    # Previous month for trend
    prev_month_start = (datetime.now(timezone.utc).replace(day=1) - timedelta(days=1)).replace(day=1).date().isoformat()
    prev_month_end = datetime.now(timezone.utc).replace(day=1).date().isoformat()
    prev_pipeline = [
        {"$match": {"tenant_id": p.tenant_id, "day": {"$gte": prev_month_start, "$lt": prev_month_end}}},
        {"$group": {"_id": None, "cost": {"$sum": "$cost_amount"}}},
    ]
    prev_data = [d async for d in db.usage_daily_rollup.aggregate(prev_pipeline)]
    prev_cost = (prev_data[0]["cost"] if prev_data else 0)

    mtd_cost = mtd.get("cost", 0)
    trend_pct = ((mtd_cost - prev_cost) / prev_cost * 100) if prev_cost else 0.0

    return {
        "window_days": days,
        "from_date": since_day,
        "to_date": to_day,
        "total_cost_usd": round(total_cost, 4),
        "total_delivered": total_delivered,
        "by_category": by_category,
        "by_country": by_country,
        "mtd_cost_usd": round(mtd_cost, 4),
        "mtd_delivered": mtd.get("delivered", 0),
        "prev_month_cost_usd": round(prev_cost, 4),
        "mtd_trend_pct": round(trend_pct, 1),
    }


# ─── Export ───────────────────────────────────────────────────────────────────

async def _build_export_response(
    db, tenant_id: str, fmt: str, from_date: str | None, to_date: str | None
):
    if not from_date:
        from_date = (datetime.now(timezone.utc) - timedelta(days=30)).date().isoformat()
    if not to_date:
        to_date = datetime.now(timezone.utc).date().isoformat()

    # Fetch rollup data
    usage_docs = await db.usage_daily_rollup.find(
        {"tenant_id": tenant_id, "day": {"$gte": from_date, "$lte": to_date}}
    ).sort("day", 1).to_list(10000)

    # Fetch template performance
    msg_pipeline = [
        {"$match": {
            "tenant_id": tenant_id,
            "direction": "outbound",
            "created_at": {"$gte": from_date, "$lte": to_date + "T23:59:59"},
            "template_name": {"$ne": None},
        }},
        {"$group": {
            "_id": "$template_name",
            "sent": {"$sum": 1},
            "delivered": {"$sum": {"$cond": [{"$in": ["$status", ["delivered", "read"]]}, 1, 0]}},
            "read": {"$sum": {"$cond": [{"$eq": ["$status", "read"]}, 1, 0]}},
            "failed": {"$sum": {"$cond": [{"$eq": ["$status", "failed"]}, 1, 0]}},
        }},
        {"$sort": {"sent": -1}},
    ]
    template_docs = [d async for d in db.messages.aggregate(msg_pipeline)]

    if fmt == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()

        # Sheet 1: Usage Rollup
        ws1 = wb.active
        ws1.title = "Usage"
        headers = ["Date", "Category", "Country", "Delivered", "Billable", "Free", "Est. Cost (USD)"]
        ws1.append(headers)
        for cell in ws1[1]:
            cell.font = Font(bold=True)
        for d in usage_docs:
            ws1.append([
                d["day"], d.get("category", ""), d.get("country_code", ""),
                d.get("delivered_count", 0), d.get("billable_count", 0),
                d.get("free_count", 0), round(d.get("cost_amount", 0.0), 4),
            ])

        # Sheet 2: Template Performance
        ws2 = wb.create_sheet("Template Performance")
        ws2.append(["Template Name", "Sent", "Delivered", "Read", "Failed", "Delivery Rate", "Read Rate"])
        for cell in ws2[1]:
            cell.font = Font(bold=True)
        for t in template_docs:
            sent = t["sent"] or 1
            ws2.append([
                t["_id"], t["sent"], t["delivered"], t["read"], t["failed"],
                f"{t['delivered'] / sent * 100:.1f}%", f"{t['read'] / sent * 100:.1f}%",
            ])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"usage_{tenant_id[:8]}_{from_date}_{to_date}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    else:
        # CSV
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Date", "Category", "Country", "Delivered", "Billable", "Free", "Est. Cost (USD)"])
        for d in usage_docs:
            writer.writerow([
                d["day"], d.get("category", ""), d.get("country_code", ""),
                d.get("delivered_count", 0), d.get("billable_count", 0),
                d.get("free_count", 0), round(d.get("cost_amount", 0.0), 4),
            ])
        writer.writerow([])
        writer.writerow(["Template Performance"])
        writer.writerow(["Template Name", "Sent", "Delivered", "Read", "Failed"])
        for t in template_docs:
            writer.writerow([t["_id"], t["sent"], t["delivered"], t["read"], t["failed"]])

        output.seek(0)
        filename = f"usage_{tenant_id[:8]}_{from_date}_{to_date}.csv"
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )


@router.get("/export")
async def export_usage(
    p: Principal = Depends(require_tenant),
    format: str = Query("csv"),
    from_date: str | None = Query(default=None),
    to_date: str | None = Query(default=None),
):
    """Download usage data as CSV or Excel (.xlsx)."""
    db = get_db()
    return await _build_export_response(db, p.tenant_id, format, from_date, to_date)
